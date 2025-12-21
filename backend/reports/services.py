"""
Reports services for TexPro AI
PDF and Excel generation with comprehensive filtering
"""

import io
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg
from django.http import HttpResponse

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def parse_date_filters(request):
    """Parse date filter parameters from request"""
    filters = {}
    
    # Date range filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    date_range = request.GET.get('date_range')
    
    if start_date:
        try:
            filters['start_date'] = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if end_date:
        try:
            filters['end_date'] = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Predefined date ranges
    if date_range:
        today = timezone.now().date()
        if date_range == 'last_7_days':
            filters['start_date'] = today - timedelta(days=7)
            filters['end_date'] = today
        elif date_range == 'last_30_days':
            filters['start_date'] = today - timedelta(days=30)
            filters['end_date'] = today
        elif date_range == 'last_quarter':
            filters['start_date'] = today - timedelta(days=90)
            filters['end_date'] = today
    
    return filters


def apply_workflow_filters(queryset, request):
    """Apply filters to workflow queryset"""
    filters = parse_date_filters(request)
    
    # Date filters
    if 'start_date' in filters:
        queryset = queryset.filter(start_date__gte=filters['start_date'])
    if 'end_date' in filters:
        queryset = queryset.filter(end_date__lte=filters['end_date'])
    
    # Status filter
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)
    
    # Product type filter
    product_type = request.GET.get('product_type')
    if product_type:
        queryset = queryset.filter(product_type=product_type)
    
    # Batch number filter
    batch_number = request.GET.get('batch_number')
    if batch_number:
        queryset = queryset.filter(batch_number__icontains=batch_number)
    
    return queryset


def apply_machine_filters(queryset, request):
    """Apply filters to machine queryset"""
    # Status filter
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)
    
    # Machine type filter
    machine_type = request.GET.get('machine_type')
    if machine_type:
        queryset = queryset.filter(machine_type=machine_type)
    
    # Location filter
    location = request.GET.get('location')
    if location:
        queryset = queryset.filter(location__icontains=location)
    
    return queryset


def apply_maintenance_filters(queryset, request):
    """Apply filters to maintenance queryset"""
    filters = parse_date_filters(request)
    
    # Date filters
    # Map generic start/end filters to model fields: MaintenanceLog uses
    # `reported_at` (when reported) and `resolved_at` (when completed).
    if 'start_date' in filters:
        queryset = queryset.filter(reported_at__date__gte=filters['start_date'])
    if 'end_date' in filters:
        queryset = queryset.filter(
            Q(resolved_at__date__lte=filters['end_date']) |
            Q(resolved_at__isnull=True)
        )
    
    # Priority filter
    priority = request.GET.get('priority')
    if priority:
        queryset = queryset.filter(priority=priority)
    
    # Status filter
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)
    
    # Maintenance type filter
    # Some clients may send `maintenance_type` but the model does not
    # have a `maintenance_type` field. Ignore this filter to avoid
    # runtime errors; if you have a specific mapping, update here.
    # maintenance_type = request.GET.get('maintenance_type')
    # if maintenance_type:
    #     queryset = queryset.filter(maintenance_type=maintenance_type)
    
    return queryset


def apply_quality_filters(queryset, request):
    """Apply filters to quality queryset"""
    filters = parse_date_filters(request)
    
    # Date filters
    # Quality checks use `created_at` as the timestamp
    if 'start_date' in filters:
        queryset = queryset.filter(created_at__date__gte=filters['start_date'])
    if 'end_date' in filters:
        queryset = queryset.filter(created_at__date__lte=filters['end_date'])
    
    # Status filter
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)
    
    # Inspector filter
    inspector = request.GET.get('inspector')
    if inspector:
        queryset = queryset.filter(inspector__username=inspector)
    
    # Defect type filter
    defect_type = request.GET.get('defect_type')
    if defect_type:
        queryset = queryset.filter(defects__icontains=defect_type)
    
    return queryset


def apply_allocation_filters(queryset, request, model_type='workforce'):
    """Apply filters to allocation queryset"""
    filters = parse_date_filters(request)
    
    if model_type == 'workforce':
        # Date filters
        if 'start_date' in filters:
            queryset = queryset.filter(start_date__gte=filters['start_date'])
        if 'end_date' in filters:
            queryset = queryset.filter(end_date__lte=filters['end_date'])
        
        # Role filter
        role = request.GET.get('role')
        if role:
            queryset = queryset.filter(role_assigned=role)
        
        # Batch filter
        batch = request.GET.get('batch')
        if batch:
            queryset = queryset.filter(batch__batch_number__icontains=batch)
    
    elif model_type == 'material':
        # Material type filter
        material_type = request.GET.get('material_type')
        if material_type:
            queryset = queryset.filter(material_name__icontains=material_type)
        
        # Batch filter
        batch = request.GET.get('batch')
        if batch:
            queryset = queryset.filter(batch__batch_number__icontains=batch)
    
    return queryset


def create_pdf_base(title, subtitle=""):
    """Create base PDF document with TexPro AI branding"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        alignment=1,  # Center alignment
        spaceAfter=12
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.darkgrey,
        alignment=1,
        spaceAfter=20
    )
    
    # Add header
    story.append(Paragraph("TexPro AI - Textile Manufacturing System", title_style))
    story.append(Paragraph(f"CMDT Mali - {title}", subtitle_style))
    
    if subtitle:
        story.append(Paragraph(subtitle, subtitle_style))
    
    # Add timestamp
    timestamp = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generated on: {timestamp}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    return buffer, doc, story, styles


def create_excel_base(title):
    """Create base Excel workbook with TexPro AI formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Title
    ws['A1'] = f"TexPro AI - {title}"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Timestamp
    ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True)
    
    return wb, ws, header_font, header_fill
